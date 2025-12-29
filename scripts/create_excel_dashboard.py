#!/usr/bin/env python3
"""
Create Excel Dashboard with Charts
Generates Dashboard_2026.xlsx with full visualization
Uses real data from SQLite database
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

# Import database functions
try:
    from finance_db import (get_monthly_summary, get_categories, get_active_installments,
                            get_pj_monthly_summary, get_pj_yearly_summary, get_consolidated_summary)
except ImportError:
    from scripts.finance_db import (get_monthly_summary, get_categories, get_active_installments,
                                    get_pj_monthly_summary, get_pj_yearly_summary, get_consolidated_summary)

# Paths
OUTPUT_PATH = Path(__file__).parent.parent / "projections" / "Dashboard_2026.xlsx"

# Get real data from database
def get_category_data(year: int, month: int):
    """Get category spending data from database."""
    summary = get_monthly_summary(year, month)
    return [(c["category"].title(), c["budget"], c["total"], int(c["percent"]))
            for c in summary["categories"]]


def get_monthly_installments():
    """Get installment projections per month for 2026, grouped by category and major items."""
    installments = get_active_installments()

    # Category mapping
    categories = ["alimentacao", "compras", "casa", "transporte", "saude",
                  "assinaturas", "lazer", "educacao", "taxas"]

    # Monthly projections by category
    monthly = {m: {cat: 0 for cat in categories} for m in range(1, 13)}
    monthly_total = {m: 0 for m in range(1, 13)}

    # Major items tracking for Fluxo de Caixa
    major_items = {m: {"moveis": 0, "mesa": 0, "eletros": 0, "outros": 0} for m in range(1, 13)}

    for inst in installments:
        amount = inst["installment_amount"]
        start = inst["start_date"]
        end = inst["end_date"]
        cat_name = inst.get("category_name", "compras") or "compras"
        description = inst.get("description", "").upper()

        # Parse dates
        start_year, start_month = int(start[:4]), int(start[5:7])
        end_year, end_month = int(end[:4]), int(end[5:7])

        # Calculate months active in 2026
        for month in range(1, 13):
            month_start = (2026, month)
            inst_start = (start_year, start_month)
            inst_end = (end_year, end_month)

            if inst_start <= month_start <= inst_end:
                if cat_name in categories:
                    monthly[month][cat_name] += amount
                monthly_total[month] += amount

                # Classify major items
                if "MOVEIS PLANEJADOS" in description:
                    major_items[month]["moveis"] += amount
                elif "MESA" in description and "CADEIRAS" in description:
                    major_items[month]["mesa"] += amount
                elif "ELETRODOMESTICOS" in description:
                    major_items[month]["eletros"] += amount
                else:
                    major_items[month]["outros"] += amount

    return monthly, monthly_total, major_items

# Colors
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_dashboard():
    """Create the complete Excel dashboard."""
    wb = Workbook()

    # Rename default sheet
    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"

    # Create other sheets
    ws_mensal = wb.create_sheet("Mensal")
    ws_categorias = wb.create_sheet("Categorias")
    ws_parcelamentos = wb.create_sheet("Parcelamentos")
    ws_fluxo = wb.create_sheet("Fluxo Anual")
    ws_previsoes = wb.create_sheet("Previsões ML")
    ws_dados = wb.create_sheet("Dados")

    # Populate sheets
    create_dashboard_sheet(ws_dashboard)
    create_mensal_sheet(ws_mensal)
    create_categorias_sheet(ws_categorias)
    create_parcelamentos_sheet(ws_parcelamentos)
    create_fluxo_sheet(ws_fluxo)
    create_previsoes_sheet(ws_previsoes)
    create_dados_sheet(ws_dados)

    # Save
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"✅ Dashboard criado: {OUTPUT_PATH}")


def create_dashboard_sheet(ws):
    """Create main dashboard with summary and charts for current month."""
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 5
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15

    # Title - Explicit month
    ws['A1'] = "JANEIRO 2026 - DASHBOARD"
    ws['A1'].font = Font(bold=True, size=18)
    ws.merge_cells('A1:E1')

    ws['A2'] = f"Atualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10)

    # Get data from database
    _, monthly_total, major_items = get_monthly_installments()
    jan_parcelamentos = int(monthly_total.get(1, 0))
    jan_summary = get_monthly_summary(2026, 1)
    gastos_variaveis_real = int(jan_summary.get("total_spent", 0))

    # Projeções de obra (do plano)
    OBRA_PROJETADO = 9590  # Elétrica + Piso janeiro

    # Calcular poupança
    RECEITA = 55000
    VARIAVEIS_PROJ = 16500
    saida_proj = VARIAVEIS_PROJ + jan_parcelamentos + OBRA_PROJETADO
    poupanca_proj = RECEITA - saida_proj
    saida_real = gastos_variaveis_real + 9500 + 0  # Parcela paga + Obra paga
    poupanca_real = RECEITA - saida_real

    # Section 1: RESUMO DO MÊS
    ws['A4'] = "RESUMO DO MÊS"
    ws['A4'].font = TITLE_FONT

    headers = ["Métrica", "Projetado", "Real", "Variação", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Calculate variations
    def calc_var(proj, real):
        if proj == 0:
            return "0%"
        var = int((real / proj - 1) * 100)
        return f"{var:+d}%" if var != 0 else "0%"

    def get_status(proj, real, is_expense=True):
        if is_expense:
            return "✅" if real <= proj else "⚠️"
        return "✅" if real >= proj else "⚠️"

    summary_data = [
        ["Receita", RECEITA, RECEITA, "0%", "✅"],
        ["Gastos Variáveis", VARIAVEIS_PROJ, gastos_variaveis_real, calc_var(VARIAVEIS_PROJ, gastos_variaveis_real), get_status(VARIAVEIS_PROJ, gastos_variaveis_real)],
        ["Parcelamentos", jan_parcelamentos, 9500, calc_var(jan_parcelamentos, 9500), get_status(jan_parcelamentos, 9500)],
        ["Obra", OBRA_PROJETADO, 0, "-100%", "🟡"],
        ["SAÍDA TOTAL", saida_proj, saida_real, calc_var(saida_proj, saida_real), get_status(saida_proj, saida_real)],
        ["POUPANÇA", poupanca_proj, poupanca_real, calc_var(poupanca_proj, poupanca_real), get_status(poupanca_proj, poupanca_real, is_expense=False)],
    ]

    for row_idx, row_data in enumerate(summary_data, 6):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in [2, 3]:
                cell.number_format = 'R$ #,##0'
            if col_idx == 5:
                cell.alignment = Alignment(horizontal='center')
                if value == "✅":
                    cell.fill = OK_FILL
                elif value == "⚠️":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif value == "🟡":
                    cell.fill = WARN_FILL
        # Bold for total rows
        if row_data[0] in ["SAÍDA TOTAL", "POUPANÇA"]:
            for col_idx in range(1, 6):
                ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)

    # Category summary for chart
    ws['A12'] = "GASTOS POR CATEGORIA"
    ws['A12'].font = TITLE_FONT

    cat_headers = ["Categoria", "Budget", "Real", "%"]
    for col, header in enumerate(cat_headers, 1):
        cell = ws.cell(row=13, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Get real data from database (January 2026)
    categories = get_category_data(2026, 1)

    for row_idx, row_data in enumerate(categories, 14):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in [2, 3]:
                cell.number_format = 'R$ #,##0'
            if col_idx == 4:
                cell.number_format = '0%'
                cell.alignment = Alignment(horizontal='center')

    # Bar Chart - Budget vs Real
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Budget vs Real por Categoria"
    chart1.y_axis.title = "R$"
    chart1.x_axis.title = "Categoria"

    data = Reference(ws, min_col=2, min_row=13, max_col=3, max_row=22)
    cats = Reference(ws, min_col=1, min_row=14, max_row=22)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    chart1.width = 18
    chart1.height = 10

    ws.add_chart(chart1, "F4")

    # Monthly trend data - dynamic from database
    ws['A25'] = "TENDENCIA MENSAL"
    ws['A25'].font = TITLE_FONT

    month_headers = ["Mes", "Projetado", "Real"]
    for col, header in enumerate(month_headers, 1):
        cell = ws.cell(row=26, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Get installments and calculate projections dynamically
    _, monthly_total, _ = get_monthly_installments()
    month_names = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun"]
    BUDGET_VARIAVEIS = 16500

    months = []
    for m in range(1, 7):
        # Projetado = Budget variáveis + Parcelamentos do mês
        projetado = BUDGET_VARIAVEIS + int(monthly_total.get(m, 0))
        # Real = Transações reais do mês (só janeiro tem dados por enquanto)
        if m == 1:
            summary = get_monthly_summary(2026, m)
            real = int(summary.get("total_spent", 0))
        else:
            real = 0
        months.append([month_names[m-1], projetado, real])

    for row_idx, row_data in enumerate(months, 27):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in [2, 3]:
                cell.number_format = 'R$ #,##0'

    # Line Chart - Monthly Trend
    chart2 = LineChart()
    chart2.title = "Gastos Mensais: Projetado vs Real"
    chart2.y_axis.title = "R$"
    chart2.x_axis.title = "Mes"

    data2 = Reference(ws, min_col=2, min_row=26, max_col=3, max_row=32)
    cats2 = Reference(ws, min_col=1, min_row=27, max_row=32)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.width = 18
    chart2.height = 10

    ws.add_chart(chart2, "F18")


def create_mensal_sheet(ws):
    """Create simplified 6-month view of variable expenses by category."""
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    for col in range(3, 9):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws['A1'] = "GASTOS VARIÁVEIS - VISÃO 6 MESES"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')

    ws['A2'] = "Gastos reais por categoria (não inclui parcelamentos/obra)"
    ws['A2'].font = Font(italic=True, size=10, color="666666")

    # Get category data
    cat_data = get_category_data(2026, 1)
    month_names = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun"]

    # Column headers: Categoria | Budget | Jan | Fev | Mar | Abr | Mai | Jun
    headers = ["Categoria", "Budget"] + month_names
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Data rows per category
    total_budget = 0
    monthly_totals = {m: 0 for m in range(1, 7)}

    row = 5
    for cat_name, budget, spent_jan, pct in cat_data:
        ws.cell(row=row, column=1, value=cat_name).border = THIN_BORDER
        ws.cell(row=row, column=2, value=budget).border = THIN_BORDER
        ws.cell(row=row, column=2).number_format = 'R$ #,##0'

        # Monthly spending (only January has data for now)
        for month_idx, month_num in enumerate(range(1, 7), 3):
            if month_num == 1:
                real = int(spent_jan)
            else:
                real = 0  # Future months - no data yet

            cell = ws.cell(row=row, column=month_idx, value=real)
            cell.border = THIN_BORDER
            cell.number_format = 'R$ #,##0'

            # Color code: red if over budget, yellow if > 80%
            if real > budget:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif budget > 0 and real > budget * 0.8:
                cell.fill = WARN_FILL

            monthly_totals[month_num] += real

        total_budget += budget
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=total_budget).font = Font(bold=True)
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=2).number_format = 'R$ #,##0'

    for month_idx, month_num in enumerate(range(1, 7), 3):
        cell = ws.cell(row=row, column=month_idx, value=monthly_totals[month_num])
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        cell.number_format = 'R$ #,##0'

    # % of budget row
    row += 1
    ws.cell(row=row, column=1, value="% do Budget").font = Font(italic=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value="").border = THIN_BORDER

    for month_idx, month_num in enumerate(range(1, 7), 3):
        pct = monthly_totals[month_num] / total_budget if total_budget > 0 else 0
        cell = ws.cell(row=row, column=month_idx, value=pct)
        cell.font = Font(italic=True)
        cell.border = THIN_BORDER
        cell.number_format = '0%'


def create_categorias_sheet(ws):
    """Create categories analysis sheet for current month - variable expenses only."""
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

    # Title - Explicit month
    ws['A1'] = "JANEIRO 2026 - GASTOS POR CATEGORIA"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')

    ws['A2'] = "Gastos variáveis do mês (não inclui parcelamentos/obra)"
    ws['A2'].font = Font(italic=True, size=10, color="666666")

    headers = ["Categoria", "Budget", "Gasto Real", "Disponível", "%", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Get real data from database
    cat_data = get_category_data(2026, 1)
    categories = [(name, budget, int(spent)) for name, budget, spent, pct in cat_data]

    total_budget = 0
    total_real = 0

    for row_idx, (cat, budget, real) in enumerate(categories, 5):
        disponivel = budget - real
        pct = real / budget if budget > 0 else 0
        status = "✅" if real <= budget else "⚠️"

        ws.cell(row=row_idx, column=1, value=cat).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=budget).border = THIN_BORDER
        ws.cell(row=row_idx, column=2).number_format = 'R$ #,##0'
        ws.cell(row=row_idx, column=3, value=real).border = THIN_BORDER
        ws.cell(row=row_idx, column=3).number_format = 'R$ #,##0'
        ws.cell(row=row_idx, column=4, value=disponivel).border = THIN_BORDER
        ws.cell(row=row_idx, column=4).number_format = 'R$ #,##0'
        # Color disponivel based on status
        if disponivel < 0:
            ws.cell(row=row_idx, column=4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif disponivel < budget * 0.2:
            ws.cell(row=row_idx, column=4).fill = WARN_FILL
        ws.cell(row=row_idx, column=5, value=pct).border = THIN_BORDER
        ws.cell(row=row_idx, column=5).number_format = '0%'
        ws.cell(row=row_idx, column=6, value=status).border = THIN_BORDER
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal='center')
        if status == "✅":
            ws.cell(row=row_idx, column=6).fill = OK_FILL

        total_budget += budget
        total_real += real

    # Total row
    total_row = 5 + len(categories)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    ws.cell(row=total_row, column=2, value=total_budget).font = Font(bold=True)
    ws.cell(row=total_row, column=2).border = THIN_BORDER
    ws.cell(row=total_row, column=2).number_format = 'R$ #,##0'
    ws.cell(row=total_row, column=3, value=total_real).font = Font(bold=True)
    ws.cell(row=total_row, column=3).border = THIN_BORDER
    ws.cell(row=total_row, column=3).number_format = 'R$ #,##0'
    ws.cell(row=total_row, column=4, value=total_budget - total_real).font = Font(bold=True)
    ws.cell(row=total_row, column=4).border = THIN_BORDER
    ws.cell(row=total_row, column=4).number_format = 'R$ #,##0'
    pct_total = total_real / total_budget if total_budget > 0 else 0
    ws.cell(row=total_row, column=5, value=pct_total).font = Font(bold=True)
    ws.cell(row=total_row, column=5).border = THIN_BORDER
    ws.cell(row=total_row, column=5).number_format = '0%'

    # Pie Chart
    chart = PieChart()
    chart.title = "Distribuição de Gastos - Janeiro 2026"

    data = Reference(ws, min_col=3, min_row=4, max_row=13)
    cats = Reference(ws, min_col=1, min_row=5, max_row=13)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 14
    chart.height = 10

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = False

    ws.add_chart(chart, "H4")


def create_parcelamentos_sheet(ws):
    """Create installments tracking sheet."""
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10

    ws['A1'] = "CONTROLE DE PARCELAMENTOS"
    ws['A1'].font = Font(bold=True, size=16)

    # Main installments
    ws['A3'] = "GRANDES COMPROMISSOS"
    ws['A3'].font = TITLE_FONT

    headers = ["Descricao", "Total", "Parcela", "Valor/Mes", "Pago", "Pendente", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    installments = [
        ["Moveis Planejados", 95000, "1/10", 9500, 9500, 85500, "Em dia"],
        ["Mesa + Cadeiras", 16000, "1/10", 1600, 0, 16000, "Novo"],
        ["Eletrodomesticos", 15000, "0/6", 2500, 0, 15000, "Futuro"],
    ]

    for row_idx, row_data in enumerate(installments, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in [2, 4, 5, 6]:
                cell.number_format = 'R$ #,##0'

    # Other installments
    ws['A10'] = "OUTROS PARCELAMENTOS"
    ws['A10'].font = TITLE_FONT

    headers2 = ["Descricao", "Total", "Parcela", "Valor/Mes", "Termino"]
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    others = [
        ["TOK STOK", 3600, "7/12", 300, "Jul/26"],
        ["KABUM", 2400, "4/12", 200, "Out/26"],
        ["MERCADO LIVRE", 2400, "3/12", 200, "Nov/26"],
        ["LEROY MERLIN", 2160, "5/12", 180, "Set/26"],
        ["CASAS BAHIA", 1800, "6/12", 150, "Ago/26"],
    ]

    for row_idx, row_data in enumerate(others, 12):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in [2, 4]:
                cell.number_format = 'R$ #,##0'


def create_fluxo_sheet(ws):
    """Create annual cash flow sheet with Projetado vs Real columns."""
    ws.column_dimensions['A'].width = 6
    for col in range(2, 12):
        ws.column_dimensions[get_column_letter(col)].width = 11

    ws['A1'] = "FLUXO DE CAIXA ANUAL 2026"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:K1')

    ws['A2'] = "Projetado (P) vs Real (R) - Visão consolidada do ano"
    ws['A2'].font = Font(italic=True, size=10, color="666666")

    # Headers with P/R distinction
    headers = ["Mês", "Receita", "Variáveis (P)", "Variáveis (R)", "Parcelam (P)",
               "Parcelam (R)", "Obra (P)", "Obra (R)", "Saída (P)", "Saída (R)", "Poupança (P)", "Poupança (R)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Get installment projections from database
    _, monthly_total, _ = get_monthly_installments()
    month_names = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

    RECEITA = 55000
    VARIAVEIS_P = 16500  # Budget mensal

    # Obra projections per month (from plan)
    obra_proj = {
        1: 9590, 2: 6250, 3: 7333, 4: 7333, 5: 7333, 6: 19083,
        7: 7084, 8: 7084, 9: 7000, 10: 7000, 11: 5850, 12: 0
    }

    # Real data (only January has data for now)
    real_data = {
        1: {"variaveis": 7453, "parcelamentos": 9500, "obra": 0},
        # Future months - no data yet
    }

    row = 5
    total_proj = {"variaveis": 0, "parcelamentos": 0, "obra": 0, "saida": 0, "poupanca": 0}
    total_real = {"variaveis": 0, "parcelamentos": 0, "obra": 0, "saida": 0, "poupanca": 0}

    for m in range(1, 13):
        parcelam_p = int(monthly_total.get(m, 0))
        obra_p = obra_proj.get(m, 0)

        # Real values (default to 0 if no data)
        var_r = real_data.get(m, {}).get("variaveis", 0)
        parc_r = real_data.get(m, {}).get("parcelamentos", 0)
        obra_r = real_data.get(m, {}).get("obra", 0)

        saida_p = VARIAVEIS_P + parcelam_p + obra_p
        saida_r = var_r + parc_r + obra_r if m in real_data else 0
        poup_p = RECEITA - saida_p
        poup_r = RECEITA - saida_r if m in real_data else 0

        # Accumulate totals
        total_proj["variaveis"] += VARIAVEIS_P
        total_proj["parcelamentos"] += parcelam_p
        total_proj["obra"] += obra_p
        total_proj["saida"] += saida_p
        total_proj["poupanca"] += poup_p

        if m in real_data:
            total_real["variaveis"] += var_r
            total_real["parcelamentos"] += parc_r
            total_real["obra"] += obra_r
            total_real["saida"] += saida_r
            total_real["poupanca"] += poup_r

        row_data = [month_names[m-1], RECEITA, VARIAVEIS_P, var_r, parcelam_p,
                    parc_r, obra_p, obra_r, saida_p, saida_r, poup_p, poup_r]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx > 1:
                cell.number_format = 'R$ #,##0'
            # Gray out Real columns for future months
            if col_idx in [4, 6, 8, 10, 12] and m not in real_data:
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=RECEITA * 12).font = Font(bold=True)
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=2).number_format = 'R$ #,##0'

    totals = [total_proj["variaveis"], total_real["variaveis"], total_proj["parcelamentos"],
              total_real["parcelamentos"], total_proj["obra"], total_real["obra"],
              total_proj["saida"], total_real["saida"], total_proj["poupanca"], total_real["poupanca"]]

    for col_idx, value in enumerate(totals, 3):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        cell.number_format = 'R$ #,##0'

    # Line chart for savings projection
    chart = LineChart()
    chart.title = "Poupança: Projetado vs Real"
    chart.y_axis.title = "R$"

    data = Reference(ws, min_col=11, min_row=4, max_col=12, max_row=16)
    cats = Reference(ws, min_col=1, min_row=5, max_row=16)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 14
    chart.height = 10

    ws.add_chart(chart, "N4")


def create_previsoes_sheet(ws):
    """Create ML predictions and budget optimization sheet."""
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 25

    ws['A1'] = "PREVISÕES E OTIMIZAÇÃO - Machine Learning"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')

    ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10, color="666666")

    # Try to load ML modules
    try:
        from ml.spending_predictor import SpendingPredictor
        from ml.budget_optimizer import BudgetOptimizer

        # Section 1: Predictions
        ws['A4'] = "PREVISÕES PARA O PRÓXIMO MÊS"
        ws['A4'].font = TITLE_FONT

        headers = ["Categoria", "Previsão", "Confiança", "Mín", "Máx", "Tendência"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        predictor = SpendingPredictor(lookback_months=6)
        report = predictor.generate_report()

        row = 6
        for category, pred in sorted(report['predictions'].items()):
            if category.startswith('_'):
                continue

            trend = report['trends'].get(category, {})
            trend_text = trend.get('description', '-')

            ws.cell(row=row, column=1, value=category.title()).border = THIN_BORDER
            ws.cell(row=row, column=2, value=pred.get('predicted', 0)).border = THIN_BORDER
            ws.cell(row=row, column=2).number_format = 'R$ #,##0'

            conf = pred.get('confidence', 0)
            ws.cell(row=row, column=3, value=conf).border = THIN_BORDER
            ws.cell(row=row, column=3).number_format = '0%'
            if conf >= 0.7:
                ws.cell(row=row, column=3).fill = OK_FILL
            elif conf >= 0.4:
                ws.cell(row=row, column=3).fill = WARN_FILL

            ws.cell(row=row, column=4, value=pred.get('lower_bound', 0)).border = THIN_BORDER
            ws.cell(row=row, column=4).number_format = 'R$ #,##0'
            ws.cell(row=row, column=5, value=pred.get('upper_bound', 0)).border = THIN_BORDER
            ws.cell(row=row, column=5).number_format = 'R$ #,##0'
            ws.cell(row=row, column=6, value=trend_text).border = THIN_BORDER

            row += 1

        # Section 2: Budget Optimization
        row += 2
        ws.cell(row=row, column=1, value="OTIMIZAÇÃO DE BUDGET")
        ws.cell(row=row, column=1).font = TITLE_FONT
        row += 1

        headers2 = ["Categoria", "Budget Atual", "Sugerido", "Mudança", "%", "Motivo"]
        for col, header in enumerate(headers2, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        row += 1

        optimizer = BudgetOptimizer(analysis_months=6)
        opt_report = optimizer.generate_recommendations_report()

        for rec in opt_report['recommendations'][:10]:  # Top 10
            ws.cell(row=row, column=1, value=rec['category'].title()).border = THIN_BORDER
            ws.cell(row=row, column=2, value=rec['current_budget']).border = THIN_BORDER
            ws.cell(row=row, column=2).number_format = 'R$ #,##0'
            ws.cell(row=row, column=3, value=rec['suggested_budget']).border = THIN_BORDER
            ws.cell(row=row, column=3).number_format = 'R$ #,##0'
            ws.cell(row=row, column=4, value=rec['change']).border = THIN_BORDER
            ws.cell(row=row, column=4).number_format = 'R$ #,##0'

            # Color code change
            if rec['change'] > 0:
                ws.cell(row=row, column=4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                ws.cell(row=row, column=4).fill = OK_FILL

            ws.cell(row=row, column=5, value=rec['change_pct'] / 100).border = THIN_BORDER
            ws.cell(row=row, column=5).number_format = '+0%;-0%'
            ws.cell(row=row, column=6, value=rec['reason'][:40]).border = THIN_BORDER

            row += 1

        # Section 3: Summary
        row += 2
        ws.cell(row=row, column=1, value="RESUMO")
        ws.cell(row=row, column=1).font = TITLE_FONT
        row += 1

        summary = opt_report['summary']
        ws.cell(row=row, column=1, value="Potencial de economia mensal:")
        ws.cell(row=row, column=2, value=abs(summary['net_change']))
        ws.cell(row=row, column=2).number_format = 'R$ #,##0'
        row += 1

        ws.cell(row=row, column=1, value="Economia anual estimada:")
        ws.cell(row=row, column=2, value=abs(summary['net_change'] * 12))
        ws.cell(row=row, column=2).number_format = 'R$ #,##0'
        ws.cell(row=row, column=2).font = Font(bold=True)

    except Exception as e:
        ws['A4'] = f"Erro ao carregar módulos ML: {str(e)}"
        ws['A5'] = "Execute: pip install numpy scikit-learn"
        ws['A6'] = "Os módulos de previsão requerem dados históricos (mínimo 3 meses)"


def create_dados_sheet(ws):
    """Create data sheet (hidden, for sync)."""
    ws['A1'] = "DADOS SINCRONIZADOS"
    ws['A1'].font = Font(bold=True)

    ws['A3'] = "sync_timestamp"
    ws['B3'] = datetime.now().isoformat()

    ws['A4'] = "source"
    ws['B4'] = "finance.db"

    ws['A6'] = "CATEGORIAS"
    headers = ["id", "nome", "budget", "gasto_jan"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=7, column=col, value=header)

    # Get real data from database
    cat_data = get_category_data(2026, 1)
    categories = [[i+1, name.lower(), budget, int(spent)]
                  for i, (name, budget, spent, pct) in enumerate(cat_data)]

    for row_idx, row_data in enumerate(categories, 8):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)


if __name__ == "__main__":
    create_dashboard()
