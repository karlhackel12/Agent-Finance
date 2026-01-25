#!/usr/bin/env python3
"""
Sync Excel Dashboard
Updates Dashboard_2026.xlsx with data from SQLite database.
Preserves charts and formatting while updating data cells.
"""

import sqlite3
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

# Paths
DB_PATH = Path(__file__).parent.parent / "data" / "finance.db"
EXCEL_PATH = Path(__file__).parent.parent / "projections" / "Dashboard_2026.xlsx"


def get_connection():
    """Get database connection."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def get_monthly_summary(year: int, month: int) -> dict:
    """Get monthly summary from database."""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT
            c.name as category,
            c.budget_monthly as budget,
            COALESCE(SUM(ABS(t.amount)), 0) as total
        FROM categories c
        LEFT JOIN transactions t ON t.category_id = c.id
            AND strftime('%Y', t.date) = ?
            AND strftime('%m', t.date) = ?
            AND t.type = 'expense'
        GROUP BY c.id
        ORDER BY c.name
    ''', (str(year), f"{month:02d}"))

    rows = cursor.fetchall()
    conn.close()

    return {row['category']: {'budget': row['budget'], 'total': row['total']} for row in rows}


def get_transactions_count(year: int, month: int) -> int:
    """Get transaction count for month."""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT COUNT(*) as count
        FROM transactions
        WHERE strftime('%Y', date) = ?
        AND strftime('%m', date) = ?
    ''', (str(year), f"{month:02d}"))

    result = cursor.fetchone()
    conn.close()

    return result['count'] if result else 0


def sync_dashboard(year: int = 2026, month: int = 1):
    """Sync database data to Excel dashboard."""
    if not EXCEL_PATH.exists():
        print(f"❌ Excel file not found: {EXCEL_PATH}")
        print("Run create_excel_dashboard.py first.")
        return False

    # Load data
    summary = get_monthly_summary(year, month)
    tx_count = get_transactions_count(year, month)

    # Load workbook
    wb = load_workbook(EXCEL_PATH)

    # Update Dashboard sheet
    ws = wb['Dashboard']

    # Update timestamp
    ws['A2'] = f"Atualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # Category mapping (row positions in Dashboard sheet)
    category_rows = {
        'alimentacao': 14,
        'compras': 15,
        'casa': 16,
        'transporte': 17,
        'saude': 18,
        'assinaturas': 19,
        'lazer': 20,
        'educacao': 21,
        'taxas': 22,
    }

    # Update category data
    total_budget = 0
    total_real = 0

    for cat_name, row in category_rows.items():
        if cat_name in summary:
            budget = summary[cat_name]['budget']
            real = summary[cat_name]['total']
            pct = int(real / budget * 100) if budget > 0 else 0

            ws.cell(row=row, column=2, value=budget)
            ws.cell(row=row, column=3, value=real)
            ws.cell(row=row, column=4, value=pct)

            total_budget += budget
            total_real += real

    # Update summary section (row 7 = Gastos Variaveis)
    ws.cell(row=7, column=2, value=total_budget)  # Projetado
    ws.cell(row=7, column=3, value=total_real)    # Real

    pct_change = int((total_real - total_budget) / total_budget * 100) if total_budget > 0 else 0
    ws.cell(row=7, column=4, value=f"{pct_change}%")

    # Update poupanca (row 9)
    receita = 55000
    # Buscar total de obra/esportes (excluídas)
    conn_excluded = get_connection()
    cursor_excluded = conn_excluded.cursor()
    cursor_excluded.execute('''
        SELECT COALESCE(SUM(ABS(t.amount)), 0) as total_excluded
        FROM transactions t
        JOIN categories c ON t.category_id = c.id
        WHERE c.is_excluded = 1
            AND strftime('%Y', t.date) = ?
            AND strftime('%m', t.date) = ?
            AND t.type = 'expense'
    ''', (str(year), f"{month:02d}"))
    total_excluded = cursor_excluded.fetchone()['total_excluded']
    conn_excluded.close()

    # Poupança = Receita - Variáveis - Obra/Esportes
    poupanca_real = receita - total_real - total_excluded
    ws.cell(row=9, column=3, value=poupanca_real)

    # Update Mensal sheet
    ws_mensal = wb['Mensal']
    month_col = month + 1  # Column B = Jan, C = Fev, etc.

    category_mensal_rows = {
        'alimentacao': 4,
        'compras': 5,
        'casa': 6,
        'transporte': 7,
        'saude': 8,
        'assinaturas': 9,
        'lazer': 10,
        'educacao': 11,
        'taxas': 12,
    }

    for cat_name, row in category_mensal_rows.items():
        if cat_name in summary:
            ws_mensal.cell(row=row, column=month_col, value=summary[cat_name]['total'])

    # Update Categorias sheet
    ws_cat = wb['Categorias']

    cat_rows = {
        'alimentacao': 4,
        'compras': 5,
        'casa': 6,
        'transporte': 7,
        'saude': 8,
        'assinaturas': 9,
        'lazer': 10,
        'educacao': 11,
        'taxas': 12,
    }

    for cat_name, row in cat_rows.items():
        if cat_name in summary:
            budget = summary[cat_name]['budget']
            real = summary[cat_name]['total']
            ws_cat.cell(row=row, column=3, value=real)
            ws_cat.cell(row=row, column=4, value=budget - real)
            ws_cat.cell(row=row, column=5, value=real / budget if budget > 0 else 0)

    # Update Dados sheet (sync metadata)
    ws_dados = wb['Dados']
    ws_dados['B3'] = datetime.now().isoformat()

    for cat_name, row in cat_rows.items():
        data_row = row + 4  # Offset for Dados sheet
        if cat_name in summary:
            ws_dados.cell(row=data_row, column=4, value=summary[cat_name]['total'])

    # Save
    wb.save(EXCEL_PATH)
    print(f"✅ Excel sincronizado: {EXCEL_PATH}")
    print(f"   Mes: {month:02d}/{year}")
    print(f"   Total gastos: R$ {total_real:,.2f}")
    print(f"   Transacoes: {tx_count}")

    return True


def sync_all_months(year: int = 2026):
    """Sync all months for the year."""
    print(f"Sincronizando todos os meses de {year}...")

    for month in range(1, 13):
        summary = get_monthly_summary(year, month)
        total = sum(cat['total'] for cat in summary.values())

        if total > 0:
            print(f"\n--- {month:02d}/{year} ---")
            sync_dashboard(year, month)

    print("\n✅ Sincronizacao completa!")


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        if sys.argv[1] == "all":
            year = int(sys.argv[2]) if len(sys.argv) > 2 else 2026
            sync_all_months(year)
        else:
            year = int(sys.argv[1]) if len(sys.argv) > 1 else 2026
            month = int(sys.argv[2]) if len(sys.argv) > 2 else 1
            sync_dashboard(year, month)
    else:
        # Default: sync current month
        now = datetime.now()
        sync_dashboard(2026, 1)  # January 2026
